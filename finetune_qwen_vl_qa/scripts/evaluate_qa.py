from __future__ import annotations

"""
Batch evaluation of the QA-pair model against labeled val.jsonl.

Runs inference on each document (all pages at once), compares the structured
JSON prediction against the gold label, and reports per-field accuracy.

Also exports predictions to an Excel file in the same multi-row layout as
1-2026-DANE.xlsx so you can open both side-by-side.

Usage:
    source finetune_qwen_vl_pytorch/.venv311/bin/activate

    # Zero-shot (no adapter):
    python finetune_qwen_vl_qa/scripts/evaluate_qa.py \
        --val-jsonl   finetune_qwen_vl_qa/data/val.jsonl \
        --model       Qwen/Qwen2.5-VL-7B-Instruct \
        --output      eval_qa_baseline.json \
        --output-xlsx eval_qa_baseline.xlsx

    # With fine-tuned adapter:
    python finetune_qwen_vl_qa/scripts/evaluate_qa.py \
        --val-jsonl      finetune_qwen_vl_qa/data/val.jsonl \
        --model          Qwen/Qwen2.5-VL-7B-Instruct \
        --lora-adapter   finetune_qwen_vl_qa/out/lora_run1 \
        --output         eval_qa_finetuned.json \
        --output-xlsx    eval_qa_finetuned.xlsx
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import torch
from peft import PeftModel
from tqdm import tqdm
from transformers import AutoProcessor, Qwen2_5_VLForConditionalGeneration

sys.path.insert(0, str(Path(__file__).parent))
from qa_utils import (
    build_qa_prompt,
    parse_json_response,
    run_inference_multipage,
    score_prediction,
)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(results: list[dict], output_path: Path) -> None:
    """
    Write predictions (and optionally gold labels) to an Excel file using
    the same multi-row layout as 1-2026-DANE.xlsx.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("  openpyxl not installed — skipping Excel export. pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # Header
    headers = [
        "Nr wniosku", "Sposób wypełnienia", "Flaga 7.9", "Nazwa inwestycji",
        "Adres", "Teren inwestycji", "Pow. zabudowy (całość)",
        "Szerokość elewacji", "Suma pow. nadziemnych", "Suma pow. podziemnych",
        "Wys. górnej krawędzi", "Wysokość zabudowy",
        "Ilość kond. nadziemnych", "Ilość kond. podziemnych",
        "Geometria dachu", "Media", "Field accuracy",
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for item in results:
        pred = item.get("prediction", {})
        score = item.get("score", {})
        accuracy = score.get("accuracy", 0.0)

        nr           = pred.get("nr_wniosku", "")
        sposob       = pred.get("sposob_wypelnienia", "")
        flaga        = pred.get("flaga_7_9", "")
        nazwa        = pred.get("nazwa_inwestycji", "")
        adres        = pred.get("adres", "")
        teren        = pred.get("teren_inwestycji", "")
        pow_cal      = pred.get("pow_zabudowy_calosc", "")
        budynki      = pred.get("budynki") or []
        media_list   = pred.get("media") or []

        # Make sure we have at least one row per record
        n_rows = max(len(budynki), len(media_list), 1)

        for i in range(n_rows):
            b = budynki[i] if i < len(budynki) else {}
            m = media_list[i] if i < len(media_list) else ""

            row = [
                nr           if i == 0 else "",
                sposob       if i == 0 else "",
                flaga        if i == 0 else "",
                nazwa        if i == 0 else "",
                adres        if i == 0 else "",
                teren        if i == 0 else "",
                pow_cal      if i == 0 else "",
                (f"{b.get('oznaczenie', '')}: {b.get('szerokosc_elewacji', '')}" if b else ""),
                (f"{b.get('oznaczenie', '')}: {b.get('suma_pow_nadziemnych', '')}" if b else ""),
                (f"{b.get('oznaczenie', '')}: {b.get('suma_pow_podziemnych', '')}" if b else ""),
                (f"{b.get('oznaczenie', '')}: {b.get('wys_gornej_krawedzi', '')}" if b else ""),
                (f"{b.get('oznaczenie', '')}: {b.get('wysokosc_zabudowy', '')}" if b else ""),
                (f"{b.get('oznaczenie', '')}: {b.get('ilosc_kond_nadziemnych', '')}" if b else ""),
                (f"{b.get('oznaczenie', '')}: {b.get('ilosc_kond_podziemnych', '')}" if b else ""),
                (f"{b.get('oznaczenie', '')}: {b.get('geometria_dachu', '')}" if b else ""),
                m,
                f"{accuracy:.1%}" if i == 0 else "",
            ]
            ws.append(row)

    # Auto-width (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(str(output_path))
    print(f"  Excel exported → {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Evaluate QA-pair VLM on labeled val.jsonl")
    ap.add_argument("--val-jsonl",    required=True)
    ap.add_argument("--model",        default="Qwen/Qwen2.5-VL-7B-Instruct")
    ap.add_argument("--lora-adapter", default="",
                    help="Path to LoRA adapter directory (omit for zero-shot baseline)")
    ap.add_argument("--output",       required=True, help="Output JSON path for results")
    ap.add_argument("--output-xlsx",  default="",    help="Optional Excel export path")
    ap.add_argument("--max-pixels",   type=int, default=501_760)
    ap.add_argument("--max-new-tokens", type=int, default=1024)
    args = ap.parse_args()

    device = "mps" if torch.backends.mps.is_available() else "cpu"
    print(f"Device: {device}")

    print(f"Loading: {args.model}")
    processor = AutoProcessor.from_pretrained(args.model, max_pixels=args.max_pixels)
    model = Qwen2_5_VLForConditionalGeneration.from_pretrained(
        args.model,
        torch_dtype=torch.float16,
        device_map=None,
    )
    if args.lora_adapter:
        print(f"Applying LoRA adapter: {args.lora_adapter}")
        model = PeftModel.from_pretrained(model, args.lora_adapter)
    model.to(device)
    model.eval()
    print("Model ready.")

    prompt = build_qa_prompt()
    val_path = Path(args.val_jsonl)
    records = [json.loads(l) for l in val_path.read_text().splitlines() if l.strip()]

    # Only evaluate labeled records
    labeled = [r for r in records if r.get("target_json")]
    print(f"Evaluating {len(labeled)} labeled documents …")

    results = []
    field_totals: dict[str, list[bool]] = defaultdict(list)

    for rec in tqdm(labeled):
        img_paths = [Path(p) for p in rec["image_paths"]]
        gold = rec["target_json"]
        if isinstance(gold, str):
            gold = json.loads(gold)

        raw = run_inference_multipage(
            model, processor, img_paths, prompt, device,
            max_new_tokens=args.max_new_tokens,
        )
        pred = parse_json_response(raw)
        sc   = score_prediction(pred, gold)

        for field, ok in sc["field_scores"].items():
            field_totals[field].append(ok)

        results.append({
            "pdf_stem":   rec.get("pdf_stem", ""),
            "nr_wniosku": rec.get("nr_wniosku", ""),
            "gold":       gold,
            "prediction": pred,
            "raw_output": raw,
            "score":      sc,
        })

        if device == "mps":
            torch.mps.empty_cache()

    # --- Summary ---
    overall = sum(sc["correct"] for r in results for sc in [r["score"]]) / max(
        1, sum(sc["total"] for r in results for sc in [r["score"]])
    )
    print(f"\nOverall field accuracy: {overall:.1%}")
    print("\nPer-field accuracy:")
    for field, vals in sorted(field_totals.items()):
        acc = sum(vals) / len(vals)
        bar = "█" * int(acc * 20)
        print(f"  {field:<40} {acc:5.1%}  {bar}")

    out_path = Path(args.output).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps({"overall_accuracy": overall, "results": results}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\nResults saved → {out_path}")

    if args.output_xlsx:
        export_to_excel(results, Path(args.output_xlsx).expanduser().resolve())


if __name__ == "__main__":
    main()
