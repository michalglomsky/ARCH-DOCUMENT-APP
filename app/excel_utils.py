from __future__ import annotations

"""
Excel read/write utilities for the app layer.

- parse_labels(path): read labels xlsx → dict keyed by nr_wniosku
- save_result(): append an extraction result to extracted_results.xlsx
- compare_result(): field-by-field diff between VLM output and Excel ground truth
- get_excel_schema(): return the field schema (for the UI grid)
"""

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font

BACKUP_DIR    = Path("/Users/michalglomski/Documents/ARCH-APP-Project-Files-Backup")
EXCEL_LABELS  = BACKUP_DIR / "1-2026-DANE.xlsx"
TEMPLATE_PATH = Path(__file__).parent.parent / "Project Files" / "EXTRACTED-DATA-TEMPLATE.xlsx"
OUTPUT_PATH   = Path(__file__).parent / "extracted_results.xlsx"

# ---------------------------------------------------------------------------
# Column indices (0-based) — matches 1-2026-DANE.xlsx layout
# ---------------------------------------------------------------------------
COL_NR, COL_SPOSOB, COL_FLAGA = 0, 1, 2
COL_NAZWA, COL_ADRES, COL_TEREN, COL_POW_CAL = 3, 4, 5, 6
COL_SZEROK, COL_POW_NAD, COL_POW_POD = 7, 8, 9
COL_WYS_KRAW, COL_WYS_ZAB = 10, 11
COL_KOND_NAD, COL_KOND_POD, COL_DACH, COL_MEDIA = 12, 13, 14, 15


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _split(cell: str) -> tuple[str, str]:
    m = re.match(r'^([^:]+?):\s*(.*)$', cell)
    return (m.group(1).strip(), m.group(2).strip()) if m else ("", cell)


def _is_header(row) -> bool:
    return _s(row[COL_NR]).lower() == "nr wniosku"


# ---------------------------------------------------------------------------
# Parse labels Excel
# ---------------------------------------------------------------------------

def parse_labels(excel_path: Path | None = None) -> dict[str, dict]:
    """Parse a labels Excel file into {nr_wniosku: record}."""
    path = excel_path or EXCEL_LABELS
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    records: dict[str, dict] = {}
    current = None

    for row in ws.iter_rows(values_only=True):
        if all(v is None for v in row) or _is_header(row):
            continue
        nr_raw = row[COL_NR]
        if nr_raw is not None and _s(nr_raw) not in ("", "None"):
            try:
                nr = str(int(float(_s(nr_raw))))
            except ValueError:
                continue
            current = {
                "nr_wniosku":        nr,
                "sposob_wypelnienia": _s(row[COL_SPOSOB]),
                "flaga_7_9":          _s(row[COL_FLAGA]),
                "nazwa_inwestycji":   _s(row[COL_NAZWA]),
                "adres":              _s(row[COL_ADRES]),
                "teren_inwestycji":   _s(row[COL_TEREN]),
                "pow_zabudowy_calosc": _s(row[COL_POW_CAL]),
                "budynki": [],
                "media":   [],
            }
            records[nr] = current
            _add_building(current, row)
            m = _s(row[COL_MEDIA])
            if m:
                current["media"].append(m)
        elif current is not None:
            _add_building(current, row)
            m = _s(row[COL_MEDIA])
            if m:
                current["media"].append(m)

    wb.close()
    return records


def _add_building(record: dict, row) -> None:
    raw = _s(row[COL_SZEROK])
    if not raw:
        return
    label, val = _split(raw)

    def v(col):
        _, x = _split(_s(row[col]))
        return x

    record["budynki"].append({
        "oznaczenie":            label or raw,
        "szerokosc_elewacji":    val or raw,
        "suma_pow_nadziemnych":  v(COL_POW_NAD),
        "suma_pow_podziemnych":  v(COL_POW_POD),
        "wys_gornej_krawedzi":   v(COL_WYS_KRAW),
        "wysokosc_zabudowy":     v(COL_WYS_ZAB),
        "ilosc_kond_nadziemnych": v(COL_KOND_NAD),
        "ilosc_kond_podziemnych": v(COL_KOND_POD),
        "geometria_dachu":       v(COL_DACH),
    })


# ---------------------------------------------------------------------------
# Schema — field definitions for the UI grid
# ---------------------------------------------------------------------------

FLAT_FIELDS = [
    ("nr_wniosku",          "Nr wniosku"),
    ("sposob_wypelnienia",  "Sposób wypełnienia"),
    ("flaga_7_9",           "Flaga 7.9"),
    ("nazwa_inwestycji",    "Nazwa inwestycji"),
    ("adres",               "Adres"),
    ("teren_inwestycji",    "Teren inwestycji"),
    ("pow_zabudowy_calosc", "Pow. zabudowy (całość)"),
]

BUILDING_FIELDS = [
    ("oznaczenie",             "Oznaczenie"),
    ("szerokosc_elewacji",     "Szerokość elewacji"),
    ("suma_pow_nadziemnych",   "Suma pow. nadziemnych"),
    ("suma_pow_podziemnych",   "Suma pow. podziemnych"),
    ("wys_gornej_krawedzi",    "Wys. górnej krawędzi"),
    ("wysokosc_zabudowy",      "Wysokość zabudowy"),
    ("ilosc_kond_nadziemnych", "Kond. nadziemne"),
    ("ilosc_kond_podziemnych", "Kond. podziemne"),
    ("geometria_dachu",        "Geometria dachu"),
]


def get_excel_schema() -> dict:
    """Return field schema used by the UI to render the editable grid."""
    return {
        "flat":     [{"key": k, "label": l} for k, l in FLAT_FIELDS],
        "building": [{"key": k, "label": l} for k, l in BUILDING_FIELDS],
    }


# ---------------------------------------------------------------------------
# Save result to Excel
# ---------------------------------------------------------------------------

def save_result(pred: dict) -> str:
    """Append a prediction to the output Excel file. Returns the output path."""
    if OUTPUT_PATH.exists():
        wb = openpyxl.load_workbook(str(OUTPUT_PATH))
        ws = wb.active
    elif TEMPLATE_PATH.exists():
        wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [l for _, l in FLAT_FIELDS] + [l for _, l in BUILDING_FIELDS[1:]] + ["Media"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    nr      = pred.get("nr_wniosku", "")
    sposob  = pred.get("sposob_wypelnienia", "")
    flaga   = pred.get("flaga_7_9", "")
    nazwa   = pred.get("nazwa_inwestycji", "")
    adres   = pred.get("adres", "")
    teren   = pred.get("teren_inwestycji", "")
    pow_cal = pred.get("pow_zabudowy_calosc", "")
    budynki = pred.get("budynki") or []
    media   = pred.get("media") or []

    n_rows = max(len(budynki), len(media), 1)
    for i in range(n_rows):
        b  = budynki[i] if i < len(budynki) else {}
        m  = media[i]   if i < len(media)   else ""
        oz = b.get("oznaczenie", "") if b else ""
        ws.append([
            nr      if i == 0 else "",
            sposob  if i == 0 else "",
            flaga   if i == 0 else "",
            nazwa   if i == 0 else "",
            adres   if i == 0 else "",
            teren   if i == 0 else "",
            pow_cal if i == 0 else "",
            (f"{oz}: {b.get('szerokosc_elewacji','')}"     if b else ""),
            (f"{oz}: {b.get('suma_pow_nadziemnych','')}"   if b else ""),
            (f"{oz}: {b.get('suma_pow_podziemnych','')}"   if b else ""),
            (f"{oz}: {b.get('wys_gornej_krawedzi','')}"    if b else ""),
            (f"{oz}: {b.get('wysokosc_zabudowy','')}"      if b else ""),
            (f"{oz}: {b.get('ilosc_kond_nadziemnych','')}" if b else ""),
            (f"{oz}: {b.get('ilosc_kond_podziemnych','')}" if b else ""),
            (f"{oz}: {b.get('geometria_dachu','')}"        if b else ""),
            m,
        ])

    wb.save(str(OUTPUT_PATH))
    return str(OUTPUT_PATH)


# ---------------------------------------------------------------------------
# Compare VLM prediction against Excel ground truth
# ---------------------------------------------------------------------------

def _norm(v: Any) -> str:
    return "" if v is None else str(v).strip().lower()


def compare_result(pred: dict, gold: dict) -> dict:
    """
    Returns a comparison dict:
    {
      "flat": { "adres": {"pred": "...", "gold": "...", "match": true}, ... },
      "budynki": [ { "szerokosc_elewacji": {...}, ... }, ... ],
      "media": { "pred": [...], "gold": [...], "match": true },
      "overall_accuracy": 0.85
    }
    """
    result: dict[str, Any] = {"flat": {}, "budynki": [], "media": {}}
    correct = total = 0

    for key, _ in FLAT_FIELDS:
        p, g = _norm(pred.get(key)), _norm(gold.get(key))
        match = p == g
        result["flat"][key] = {"pred": pred.get(key, ""), "gold": gold.get(key, ""), "match": match}
        correct += int(match); total += 1

    pred_blds = pred.get("budynki") or []
    gold_blds = gold.get("budynki") or []
    for i in range(max(len(pred_blds), len(gold_blds))):
        pb = pred_blds[i] if i < len(pred_blds) else {}
        gb = gold_blds[i] if i < len(gold_blds) else {}
        bld_cmp = {}
        for key, _ in BUILDING_FIELDS:
            p, g = _norm(pb.get(key)), _norm(gb.get(key))
            match = p == g
            bld_cmp[key] = {"pred": pb.get(key, ""), "gold": gb.get(key, ""), "match": match}
            correct += int(match); total += 1
        result["budynki"].append(bld_cmp)

    pred_media = set(_norm(m) for m in (pred.get("media") or []))
    gold_media = set(_norm(m) for m in (gold.get("media") or []))
    match = pred_media == gold_media
    result["media"] = {
        "pred": pred.get("media") or [],
        "gold": gold.get("media") or [],
        "match": match,
    }
    correct += int(match); total += 1

    result["overall_accuracy"] = round(correct / max(1, total), 3)
    return result
